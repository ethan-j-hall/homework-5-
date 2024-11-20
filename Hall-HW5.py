#Ethan Hall
#COSC 1010
#11.19.2024
#Sources: Class powerpoints, W3 Schools, Google AI suggestion box 

import openpyxl
from openpyxl.styles import PatternFill

pixel_art = [
    ['red', 'red', 'red', 'blue', 'blue'],
    ['red', 'white', 'red', 'blue', 'green'],
    ['red', 'red', 'white', 'green', 'green'],
    ['black', 'black', 'black', 'green', 'green'],
    ['black', 'black', 'black', 'green', 'green']
]

# Color map
color_map = {
    'white': 'FFFFFF',
    'black': '000000',
    'red': 'FF0000',
    'blue': '0000FF',
    'green': '00FF00'
}


wb = openpyxl.Workbook()
ws = wb.active

column_width = 3
row_height = column_width * 6
ws.column_dimensions['A'].width = column_width
ws.row_dimensions[1].height = row_height

def fill_pixel(x, y, color):
    fill = PatternFill(start_color=color_map[color], end_color=color_map[color], fill_type="solid")
    ws.cell(row=y + 1, column=x + 1).fill = fill

for row_idx, row in enumerate(pixel_art):
    for col_idx, color in enumerate(row):
        fill_pixel(col_idx, row_idx, color)

wb.save('pixel_art.xlsx')
